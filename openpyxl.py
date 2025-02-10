import openpyxl

# Load the workbook
wb = openpyxl.load_workbook("example.xlsx")

# Select the worksheet
ws = wb["Sheet1"]

# Find the last row with data
last_row = ws.max_row  # Get the last row after adding data

# Locate the chart in the worksheet
for drawing in ws._charts:
    chart = drawing  # Assuming there's only one chart; if multiple, modify accordingly

    # Update the chart's data range (Assuming it's based on column A and B)
    new_data_range = f"Sheet1!$A$2:$A${last_row},Sheet1!$B$2:$B${last_row}"
    
    # Update the chart series with the new range
    chart.series[0].values = new_data_range  # Assuming it's the first series

# Save the workbook with the updated chart range
wb.save("example.xlsx")

print("Chart updated successfully!")

############################################################################################################################

# Update a formula in a specific cell (e.g., C2)
ws["C2"].value = "=SUM(A2:B2)"  # Updating the formula

############################################################################################################################

# Find the last row with data
last_row = ws.max_row  # Get the last row with data

# Update the formula in the cell (assuming it's in B6, update accordingly)
ws["B{}".format(last_row)].value = f"=SUM(A2:A{last_row})"